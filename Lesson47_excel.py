import openpyxl

book = openpyxl.open('my_excel.xlsx', read_only=True)
sheet = book.active#1 list

print(sheet['B2'].value)#получить значение ячейки 
print(sheet[1][2].value)#получить значение ячейки: строка 1 столбец 3(С)

for row in range(1,sheet.max_row+1):
    author = sheet[row][0].value
    name = sheet[row][1].value
    year = sheet[row][2].value
    rating = sheet[row][3].value
    print(row,  author,name,year,rating)

cells = sheet['B2':'C3']
for name, year in cells:
    print(name.value, year.value)

for row in sheet.iter_rows(min_row=2, max_row=4, min_col=1, max_col=3):
    for cell in row:
        print(cell.value, end= " ")
    print()
book.close()

book = openpyxl.Workbook()
sheet = book.active
sheet['A5'] = 100
sheet['B5'] = 'hello'
sheet[1][0].value = 'world'#A1
sheet.cell(row=5,column=3).value = 'hello world'#C5

book.save('my_excel1.xlsx')
book.close()

import json
with open('movies.json') as file:
    data = json.load(file)
for movie in data['movies']:
    id = movie['id']
    title = movie['title']
    year = movie['year']
    genres = movie['genres']
    director = movie['director']
    actors = movie['actors']
    print(id, title, year, genres, director, actors)

book = openpyxl.Workbook()
sheet = book.active
sheet['A1'] = 'ID'
sheet['B1'] = 'TITLE'
sheet['C1'] = 'YEAR'
sheet['D1'] = 'GENRES'
sheet['E1'] = 'DIRECTOR'
sheet['F1'] = 'ACTORS'

row = 2
for movie in data['movies']:
    sheet[row][0].value = movie['id']
    sheet[row][1].value = movie['title']
    sheet[row][2].value = movie['year']
    sheet[row][3].value = ' '.join(movie['genres'])
    sheet[row][4].value = movie['director']
    sheet[row][5].value = movie['actors']
    row += 1
book.save('my_excel1.xlsx')
book.close()